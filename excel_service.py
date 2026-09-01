"""
excel_service.py

Builds a formatted Excel (.xlsx) file from the confirmed attendance data,
using the teacher's current table configuration (columns + number of weeks).
"""

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_attendance_excel(students, config):
    columns = config.get("columns", [])
    num_weeks = config.get("numWeeks", 14)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns) + num_weeks)
    title_cell = ws.cell(row=1, column=1, value="Attendance Register")
    title_cell.font = title_font
    title_cell.alignment = center_align

    # Date generated row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns) + num_weeks)
    date_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    date_cell.alignment = center_align
    date_cell.font = Font(italic=True, size=9, color="6B7280")

    header_row_num = 4

    # Column headers
    header_labels = [c["label"] for c in columns] + [f"Wk{w}" for w in range(1, num_weeks + 1)]
    for col_index, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=header_row_num, column=col_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Student rows
    for row_offset, student in enumerate(students):
        row_num = header_row_num + 1 + row_offset
        weekly_attendance = student.get("weekly_attendance", [])
        status_by_week = {entry["week"]: entry["status"] for entry in weekly_attendance}

        col_index = 1
        for col in columns:
            cell = ws.cell(row=row_num, column=col_index, value=student.get(col["id"], ""))
            cell.border = thin_border
            col_index += 1

        for week in range(1, num_weeks + 1):
            status = status_by_week.get(week, "")
            cell = ws.cell(row=row_num, column=col_index, value=status)
            cell.alignment = center_align
            cell.border = thin_border
            if status == "Present":
                cell.font = Font(color="065F46")
            elif status == "Absent":
                cell.font = Font(color="991B1B")
            elif status == "Late":
                cell.font = Font(color="92400E")
            col_index += 1

    # Auto-adjust column widths based on content length
    total_columns = len(columns) + num_weeks
    for col_index in range(1, total_columns + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=header_row_num, max_col=col_index, min_col=col_index):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_index)].width = max(max_length + 2, 8)

    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()

    return excel_bytes